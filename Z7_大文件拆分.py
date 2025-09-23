import os
import pandas as pd
from openpyxl import load_workbook
from pandas import ExcelWriter
from config import AREA, AREA_FILE


def split_large_excel(
    source_path, 
    dest_dir, 
    max_rows=50000,
    sheet_name="数据",
    sync_columns=[2,3,4,5,6,7]  # 新增：同步A列合并格式的列（默认B-G列，1=A,2=B...）
):
    """
    拆分大型Excel文件：
    1. 合并单元格的行不跨文件；
    2. A列（第1列）合并格式与原文件一致；
    3. 可通过sync_columns参数指定需要与A列同步合并的列；
    4. 统一Sheet名称。
    
    参数:
    sync_columns: 列索引列表（1-based），指定需要与A列同步合并的列
                  例如：[2,3,4]表示B、C、D列与A列同步合并
    """
    os.makedirs(dest_dir, exist_ok=True)

    # -------------------------- 第一步：读取原文件A列合并信息与总行数 --------------------------
    wb_source = load_workbook(source_path, data_only=True)
    sheet_source = wb_source.active
    total_excel_rows = sheet_source.max_row  # 原文件总行数（Excel行号从1开始）
    merged_ranges = []  # 存储A列合并信息：{start, end, value}

    # 仅读取A列（第1列）的合并范围
    if sheet_source.merged_cells:
        for merged_range in sheet_source.merged_cells.ranges:
            if merged_range.min_col == 1 and merged_range.max_col == 1:  # 只筛选A列合并
                merged_ranges.append({
                    "start": merged_range.min_row,    # 原文件A列合并起始行
                    "end": merged_range.max_row,      # 原文件A列合并结束行
                    "value": sheet_source.cell(row=merged_range.min_row, column=1).value
                })
    wb_source.close()
    merged_ranges.sort(key=lambda x: x["start"])  # 按起始行排序

    # -------------------------- 第二步：生成不可拆分的块（合并块+普通块） --------------------------
    blocks = []
    current_row = 1  # 从原文件第1行开始遍历

    # 插入普通块和合并块（基于A列合并范围）
    for merge in merged_ranges:
        merge_start, merge_end = merge["start"], merge["end"]
        if current_row < merge_start:
            blocks.append({
                "type": "normal", "start": current_row, "end": merge_start - 1,
                "row_count": merge_start - current_row
            })
        blocks.append({
            "type": "merged", "start": merge_start, "end": merge_end,
            "row_count": merge_end - merge_start + 1, "value": merge["value"]
        })
        current_row = merge_end + 1

    if current_row <= total_excel_rows:
        blocks.append({
            "type": "normal", "start": current_row, "end": total_excel_rows,
            "row_count": total_excel_rows - current_row + 1
        })

    # -------------------------- 第三步：计算拆分断点（确保合并块完整） --------------------------
    split_points = []
    current_file_rows = 0
    current_file_start = 1

    for block in blocks:
        block_row_count = block["row_count"]
        if current_file_rows + block_row_count > max_rows:
            split_points.append((current_file_start, block["start"] - 1))
            current_file_start = block["start"]
            current_file_rows = block_row_count
        else:
            current_file_rows += block_row_count
    split_points.append((current_file_start, total_excel_rows))

    # -------------------------- 第四步：生成文件+同步合并列 --------------------------
    print(f"原文件总行数: {total_excel_rows}，预计拆分 {len(split_points)} 个文件（Sheet名：{sheet_name}）")
    print(f"与A列同步合并的列：{sync_columns}（1=A,2=B...）")
    
    for file_idx, (excel_start, excel_end) in enumerate(split_points, 1):
        # 1. 读取当前文件数据
        rows_to_read = excel_end - excel_start + 1
        skip_rows = excel_start - 1
        df = pd.read_excel(source_path, header=None, skiprows=skip_rows, nrows=rows_to_read)

        # 2. 保存基础Excel文件
        file_name = f"split_{file_idx}.xlsx"
        file_path = os.path.join(dest_dir, file_name)
        with ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

        # 3. 还原合并格式（A列+指定同步列）
        wb_new = load_workbook(file_path)
        sheet_new = wb_new[sheet_name]

        # 筛选当前文件包含的A列合并块
        merged_in_current_file = [
            merge for merge in merged_ranges
            if merge["start"] >= excel_start and merge["end"] <= excel_end
        ]

        # 处理A列合并（还原原文件格式）
        for merge in merged_in_current_file:
            new_start_row = merge["start"] - excel_start + 1
            new_end_row = merge["end"] - excel_start + 1
            sheet_new.merge_cells(
                start_row=new_start_row, start_column=1,
                end_row=new_end_row, end_column=1
            )

        # 处理同步列合并（与A列相同范围）
        for col in sync_columns:
            for merge in merged_in_current_file:
                new_start_row = merge["start"] - excel_start + 1
                new_end_row = merge["end"] - excel_start + 1
                sheet_new.merge_cells(
                    start_row=new_start_row, start_column=col,
                    end_row=new_end_row, end_column=col
                )

        # 4. 保存文件
        wb_new.save(file_path)
        wb_new.close()

        # 5. 打印进度
        print(f"已生成文件: {file_path} | 行数：{len(df)} | 合并块数：{len(merged_in_current_file)}")

    print(f"\n拆分完成！共生成 {len(split_points)} 个文件，指定列已与A列同步合并。")

if __name__ == "__main__":

    print("正在执行死值！！！！！")
    # 实际路径配置
    source_file_path = f'/Users/liuhaojun/Documents/项目文档/中国华电项目(云南)/03 时序数据质量稽核规则/{AREA_FILE}/00 总体/{AREA}区域_时序稽核质量规则_死值.xlsx'
    destination_directory = f'/Users/liuhaojun/Documents/项目文档/中国华电项目(云南)/03 时序数据质量稽核规则/{AREA_FILE}/01 死值'
    
    # 调用示例：指定B-F列（2-6）与A列同步合并
    split_large_excel(
        source_path=source_file_path,
        dest_dir=destination_directory,
        max_rows=150000,
        sheet_name="死值",  # 建议与文件内容匹配
        sync_columns=[2,3,4,5,6,7]  # B-F列（2=B,6=F）
    )

    print("正在执行跳变！！！！！")
    # 实际路径配置
    source_file_path1 = f'/Users/liuhaojun/Documents/项目文档/中国华电项目(云南)/03 时序数据质量稽核规则/{AREA_FILE}/00 总体/{AREA}区域_时序稽核质量规则_跳变.xlsx'
    destination_directory1 = f'/Users/liuhaojun/Documents/项目文档/中国华电项目(云南)/03 时序数据质量稽核规则/{AREA_FILE}/02 跳变'
    
    # 调用示例：指定B-F列（2-6）与A列同步合并
    split_large_excel(
        source_path=source_file_path1,
        dest_dir=destination_directory1,
        max_rows=150000,
        sheet_name="跳变",  # 建议与文件内容匹配
        sync_columns=[2,3,4,5,6]  # B-F列（2=B,6=F）
    )


    print("正在执行中断！！！！！")
    # 实际路径配置
    source_file_path2 = f'/Users/liuhaojun/Documents/项目文档/中国华电项目(云南)/03 时序数据质量稽核规则/{AREA_FILE}/00 总体/{AREA}区域_时序稽核质量规则_中断.xlsx'
    destination_directory2 = f'/Users/liuhaojun/Documents/项目文档/中国华电项目(云南)/03 时序数据质量稽核规则/{AREA_FILE}/04 中断'
    
    # 调用示例：指定B-F列（2-6）与A列同步合并
    split_large_excel(
        source_path=source_file_path2,
        dest_dir=destination_directory2,
        max_rows=150000,
        sheet_name="中断",  # 建议与文件内容匹配
        sync_columns=[2,3,4,5]  # B-F列（2=B,6=F）
    )
    