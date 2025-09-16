import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import psycopg2
from config import AREA
from config import AREA_FILE



def merge_columns_hierarchical(df, output_file, columns_to_merge, custom_headers=None):
    """
    按列顺序层级合并单元格（树形结构），支持自定义表头
    参数新增：
    custom_headers: 字典，键为原列名，值为自定义表头名称（如{'standard_name': '标准化名称'}）
    """
    if df is None or df.empty:
        print("无数据可处理")
        return
    
    # -------------------------- 新增：处理自定义表头 --------------------------
    # 复制原DataFrame避免修改源数据，然后应用自定义表头
    df_with_headers = df.copy()
    if custom_headers:
        # 只重命名存在的列（避免原列名不存在导致报错）
        valid_headers = {col: name for col, name in custom_headers.items() if col in df_with_headers.columns}
        df_with_headers.rename(columns=valid_headers, inplace=True)
        # 同步更新合并列名（将原合并列名替换为自定义表头名）
        columns_to_merge = [valid_headers.get(col, col) for col in columns_to_merge]
    # --------------------------------------------------------------------------
    
    # 写入带自定义表头的数据到Excel
    df_with_headers.to_excel(output_file, index=False, engine='openpyxl')
    wb = load_workbook(output_file)
    ws = wb.active
    
    # 初始区间：整个数据范围（数据从第2行开始，第1行是表头）
    current_ranges = [(2, ws.max_row)]
    
    # 按列顺序依次处理（核心：后一列的合并范围受前一列限制）
    for col_name in columns_to_merge:
        # 注意：此时col_name已替换为自定义表头名（若有），需用df_with_headers找列索引
        col_idx = df_with_headers.columns.get_loc(col_name) + 1  # Excel列索引（从1开始）
        new_ranges = []
        
        for (start_row, end_row) in current_ranges:
            if start_row > end_row:
                continue
            
            current_value = ws.cell(row=start_row, column=col_idx).value
            sub_start = start_row
            
            # 遍历当前范围内的每一行
            for row in range(start_row + 1, end_row + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                
                if cell_value != current_value:
                    # 超过1行才合并
                    if row - sub_start > 1:
                        ws.merge_cells(
                            start_row=sub_start,
                            start_column=col_idx,
                            end_row=row - 1,
                            end_column=col_idx
                        )
                        ws.cell(row=sub_start, column=col_idx).alignment = Alignment(vertical='center')
                    
                    new_ranges.append((sub_start, row - 1))
                    current_value = cell_value
                    sub_start = row
            
            # 处理当前范围的最后一个子区间
            if end_row - sub_start >= 0:
                if end_row - sub_start > 0:
                    ws.merge_cells(
                        start_row=sub_start,
                        start_column=col_idx,
                        end_row=end_row,
                        end_column=col_idx
                    )
                    ws.cell(row=sub_start, column=col_idx).alignment = Alignment(vertical='center')
                
                new_ranges.append((sub_start, end_row))
        
        current_ranges = new_ranges
    
    # 保存文件
    wb.save(output_file)
    print(f"层级合并完成，结果保存至: {output_file}")